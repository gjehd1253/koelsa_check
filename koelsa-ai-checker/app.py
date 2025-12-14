import streamlit as st
import pandas as pd
import openpyxl
import math
import re
import os
import difflib
from openai import OpenAI
from pathlib import Path
from datetime import datetime, date
import pickle
import numpy as np
import fitz  # pymupdf
from rapidfuzz import fuzz
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

# =========================
# 기본 설정
# =========================
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None
BASE_DIR = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "rules.xlsx"
EMBLEM_PATH = "emblem.png"

# ✅ KB source(파일명) → 실제 PDF 경로 매핑
SOURCE_TO_PDF = {
    "승강기 안전기준 연혁집.pdf": "data/승강기 안전기준 연혁집.pdf",
    "검사방법 표준화.pdf": "data/검사방법 표준화.pdf",
}

def resolve_pdf_path(src: str):
    """kb의 source(파일명 or 경로)를 실제 로컬 pdf 경로로 해석"""
    if not src:
        return None

    # 1) 딕셔너리 매핑 우선
    p = SOURCE_TO_PDF.get(src)
    if p and os.path.exists(p):
        return p

    # 2) src가 이미 경로인 경우
    if os.path.exists(src):
        return src

    # 3) 파일명만 오면 data/ 밑에서 찾기
    p2 = os.path.join("data", src)
    if os.path.exists(p2):
        return p2

    return None

def fmt_ymd(x):
    """Timestamp/문자열/날짜 모두 YYYY-MM-DD로 표시"""
    if x is None or x == "":
        return ""
    try:
        return pd.to_datetime(x).strftime("%Y-%m-%d")
    except Exception:
        return str(x).split(" ")[0]

# =========================
# 로컬 RAG 지식베이스 로드
# =========================
@st.cache_resource(show_spinner="기준 지식베이스를 불러오는 중입니다... (최초 1회만 실행)")
def get_kb():
    KB_PATH = "kb_standards.pkl"
    with open(KB_PATH, "rb") as f:
        kb = pickle.load(f)
    model = SentenceTransformer(kb["model_name"])
    return kb, model

def search_local_standard(question, top_k=4):
    try:
        kb_data, kb_model = get_kb()
    except FileNotFoundError:
        return []

    query_emb = kb_model.encode([question], convert_to_numpy=True)
    docs_emb = kb_data["embeddings"]
    sims = cosine_similarity(query_emb, docs_emb)[0]
    idxs = np.argsort(-sims)[:top_k]

    results = []
    for idx in idxs:
        chunk = kb_data["chunks"][idx]
        results.append(
            {
                "score": float(sims[idx]),
                "source": chunk.get("source", "문서"),
                "page": chunk.get("page", ""),
                "text": chunk.get("text", ""),
            }
        )
    return results

def ask_gpt_about_standard(question, context_text=""):
    if not OPENAI_API_KEY or client is None:
        return "※ GPT API 키가 설정되어 있지 않아 외부 질의는 사용할 수 없습니다."

    system_msg = """
당신은 승강기 안전기준 및 검사방법 표준화 문서를 설명하는 전문가입니다.

규칙:
1) 반드시 문단을 나누어 작성하세요.
2) 항목별로 들여쓰기와 줄바꿈을 사용하세요.
3) 판단(적합/부적합)은 하지 말고, 설명/해석/인용만 제공합니다.
4) 제공된 참고정보(PDF 발췌문) 내용만 사용하세요.

[출력 형식]
- 요지
- 근거 발췌(문서명/쪽수 포함)
- 해석/적용 시 유의점
"""
    user_msg = f"질문: {question}\n\n참고 정보:\n{context_text}"

    try:
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_msg.strip()},
                {"role": "user", "content": user_msg.strip()},
            ],
            temperature=0.2,
        )
        return resp.choices[0].message.content.strip()
    except Exception as e:
        return f"GPT API 호출 중 오류가 발생했습니다: {e}"

# =========================
# 엑셀 로딩
# =========================
@st.cache_data
def load_excel_frames():
    xls = pd.ExcelFile(EXCEL_PATH)
    df_input = pd.read_excel(xls, sheet_name="input")
    df_rules = pd.read_excel(xls, sheet_name="rules")
    df_stand = pd.read_excel(xls, sheet_name="rules_stand")
    df_stand.columns = df_stand.columns.astype(str).str.strip()
    return df_input, df_rules, df_stand

def preprocess_df_stand(df_stand: pd.DataFrame) -> pd.DataFrame:
    df = df_stand.copy()
    df.columns = [str(c).strip() for c in df.columns]
    if "항목" in df.columns:
        df["항목"] = df["항목"].astype(str).str.strip()
    for c in ["기준적용일", "기준전환일", "반환일", "기준적용일 반환일", "기준종료일"]:
        if c in df.columns:
            df[c] = df[c].apply(lambda x: x if pd.isna(x) else str(x).strip())
    return df

@st.cache_data
def load_input_structure():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    ws = wb["input"]

    cell_to_key = {}
    formula_key_to_formula = {}

    for row in ws.iter_rows(min_row=2):
        key = row[1].value      # B열: 입력키
        val_cell = row[2]       # C열: 값 셀

        if key:
            cell_to_key[val_cell.coordinate] = key

        if isinstance(val_cell.value, str) and val_cell.value.startswith("="):
            formula_key_to_formula[key] = val_cell.value

    return cell_to_key, formula_key_to_formula

@st.cache_data
def load_bs_values():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if "BS" not in wb.sheetnames:
        return {}
    ws = wb["BS"]
    values = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                values[cell.coordinate] = float(cell.value)
    return values

@st.cache_data
def load_bs_tables():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if "BS" not in wb.sheetnames:
        return {"payload_points": [], "area_max_points": [], "person_points": [], "area_min_points": []}

    ws = wb["BS"]
    payload_points, area_max_points, person_points, area_min_points = [], [], [], []

    for row in ws.iter_rows(min_row=2, values_only=True):
        a, b, _, d, e = row
        if all(v is None for v in row):
            break

        if isinstance(a, (int, float)):
            payload_points.append(float(a))
            area_max_points.append(float(b) if isinstance(b, (int, float)) else None)

        if isinstance(d, (int, float)):
            person_points.append(float(d))
            area_min_points.append(float(e) if isinstance(e, (int, float)) else None)

    return {
        "payload_points": payload_points,
        "area_max_points": area_max_points,
        "person_points": person_points,
        "area_min_points": area_min_points,
    }

@st.cache_data
def load_rules_formulas():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    ws = wb["rules"]

    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [c.value for c in header_cells]

    col_min = headers.index("최소값") + 1
    col_max = headers.index("최대값") + 1
    col_val = headers.index("입력값") + 1

    formulas_min, formulas_max, formulas_val = {}, {}, {}

    for row_idx in range(2, ws.max_row + 1):
        idx = row_idx - 2
        cmin = ws.cell(row=row_idx, column=col_min)
        cmax = ws.cell(row=row_idx, column=col_max)
        cval = ws.cell(row=row_idx, column=col_val)

        if isinstance(cmin.value, str) and cmin.value.startswith("="):
            formulas_min[idx] = cmin.value
        if isinstance(cmax.value, str) and cmax.value.startswith("="):
            formulas_max[idx] = cmax.value
        if isinstance(cval.value, str) and cval.value.startswith("="):
            formulas_val[idx] = cval.value

    return formulas_min, formulas_max, formulas_val

# =========================
# 계산 유틸
# =========================
ALLOWED_EXPR_CHARS = set("0123456789+-*/()., ")

def numeric_or_none(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and math.isnan(val):
            return None
        return float(val)
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        s = s.replace(",", "")
        m = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", s)
        if not m:
            return None
        try:
            return float(m.group(0))
        except ValueError:
            return None
    return None

def evaluate_excel_formula(formula, cell_to_key, input_values):
    if not isinstance(formula, str):
        return None
    expr = formula.strip()
    if expr.startswith("="):
        expr = expr[1:]

    def repl(match):
        coord = match.group(0)
        key = cell_to_key.get(coord)
        if not key:
            return "0"
        v = numeric_or_none(input_values.get(key))
        return "0" if v is None else str(v)

    expr2 = re.sub(r"[A-Z]+[0-9]+", repl, expr)
    if any(c not in ALLOWED_EXPR_CHARS for c in expr2):
        return None

    try:
        return eval(expr2, {"__builtins__": {}})
    except Exception:
        return None

def evaluate_rules_formula(formula, input_cell_to_key, input_values, bs_values):
    if not isinstance(formula, str):
        return None
    expr = formula.strip()
    if expr.startswith("="):
        expr = expr[1:]

    pattern = r"(?:'(?P<sheet1>[^']+)'!|(?P<sheet2>[A-Za-z0-9_]+)!)?(?P<coord>\$?[A-Z]+\$?[0-9]+)"

    def repl(match):
        sheet = match.group("sheet1") or match.group("sheet2")
        coord = match.group("coord").replace("$", "")

        if sheet and sheet.lower() == "input":
            key = input_cell_to_key.get(coord)
            if not key:
                return "0"
            v = numeric_or_none(input_values.get(key))
            return "0" if v is None else str(v)

        if sheet and sheet.lower() == "bs":
            v = bs_values.get(coord)
            return "0" if v is None else str(v)

        return "0"

    expr2 = re.sub(pattern, repl, expr)

    func_map = {
        "ROUND": "round",
        "ROUNDDOWN": "math.floor",
        "ROUNDUP": "math.ceil",
        "MAX": "max",
        "MIN": "min",
        "ABS": "abs",
        "INT": "int",
    }
    for excel_name, py_name in func_map.items():
        expr2 = re.sub(rf"\b{excel_name}\b", py_name, expr2, flags=re.IGNORECASE)

    expr2 = expr2.replace(";", ",")

    try:
        return eval(
            expr2,
            {
                "__builtins__": None,
                "round": round,
                "max": max,
                "min": min,
                "abs": abs,
                "int": int,
                "math": math,
            },
        )
    except Exception:
        return None

def compute_car_area_min(input_values, bs_tables):
    n = numeric_or_none(input_values.get("personnel"))
    if n is None:
        return None
    if n > 20:
        return 3.13 + 0.115 * (n - 20)

    persons = bs_tables["person_points"]
    areas = bs_tables["area_min_points"]
    for p, a in zip(persons, areas):
        if p == n:
            return a
    if persons and areas:
        idx = min(range(len(persons)), key=lambda i: abs(persons[i] - n))
        return areas[idx]
    return None

def compute_car_area_max(input_values, bs_tables):
    w = numeric_or_none(input_values.get("payload"))
    if w is None:
        return None
    if w > 2500:
        return 5 + 0.0016 * (w - 2500)

    payloads = bs_tables["payload_points"]
    areas = bs_tables["area_max_points"]
    L = min(len(payloads), len(areas))
    payloads, areas = payloads[:L], areas[:L]

    if not payloads:
        return None
    if w <= payloads[0]:
        return areas[0]

    for i in range(L - 1):
        x0, x1 = payloads[i], payloads[i + 1]
        y0, y1 = areas[i], areas[i + 1]
        if y0 is None or y1 is None:
            continue
        if x0 <= w <= x1:
            return y0 + (w - x0) * (y1 - y0) / (x1 - x0)

    for j in range(L - 1, -1, -1):
        if areas[j] is not None:
            return areas[j]
    return None

def judge_conform(value, min_val, max_val):
    v = numeric_or_none(value)
    mn = numeric_or_none(min_val)
    mx = numeric_or_none(max_val)
    if v is None:
        return "판단불가"

    def r2(x):
        return None if x is None else round(x, 2)

    v2, mn2, mx2 = r2(v), r2(mn), r2(mx)

    if mn2 is not None and mx2 is not None:
        return "적합" if (mn2 <= v2 <= mx2) else "부적합"
    if mn2 is not None:
        return "적합" if v2 >= mn2 else "부적합"
    if mx2 is not None:
        return "적합" if v2 <= mx2 else "부적합"
    return "판단불가"

def short_quote(text, max_len=600):
    t = " ".join(str(text).split())
    return t[:max_len] + ("..." if len(t) > max_len else "")

# =========================
# PDF에서 근거 캡처
# =========================
def _normalize(s: str) -> str:
    return " ".join((s or "").replace("\n", " ").split()).strip()

def split_sentences_ko(text: str):
    if not text:
        return []
    t = str(text).replace("\r", "\n")
    parts = re.split(r"(?<=[\.\?\!])\s+|\n+|;\s*", t)
    sents = []
    for p in parts:
        p = " ".join(p.split()).strip()
        if len(p) >= 6:
            sents.append(p)
    return sents

def find_best_bbox_on_page(page: fitz.Page, target_text: str, min_ratio: int = 70):
    target = _normalize(target_text)[:600]
    if not target:
        return None

    blocks = page.get_text("blocks")
    best = (0, None)

    for b in blocks:
        x0, y0, x1, y1, text, *_ = b
        cand = _normalize(text)

        if not cand:
            continue
        # 쪽수/머리말/너무 짧은 블록 제외
        if len(cand) < 40:
            continue
        if re.fullmatch(r"[\d\s\-–—]+", cand):
            continue

        r = fuzz.partial_ratio(target, cand)
        if r > best[0]:
            best = (r, fitz.Rect(x0, y0, x1, y1))

    if best[0] < min_ratio:
        return None
    return best[1]

def render_pdf_snippet(pdf_path: str, page_no_1based: int, quote: str,
                       margin: float = 12, zoom: float = 2.0,
                       pad_x_ratio: float = 0.20, pad_y_ratio: float = 0.35,
                       min_w_ratio: float = 0.85, min_h_px: float = 180):
    """
    - bbox 찾은 뒤, 자동으로 주변을 더 넓게 확장해서 '잘림' 최소화
    - 너무 가늘게 잡히면(폭이 좁으면) 페이지 폭에 가깝게 확장
    """
    doc = fitz.open(pdf_path)
    page = doc.load_page(page_no_1based - 1)
    page_rect = page.rect

    quote_clean = _normalize(quote)
    sents = split_sentences_ko(quote_clean)
    probe = sents[0] if sents else quote_clean
    probe = re.sub(r"\s*\d+\s*$", "", probe)
    probe = probe[:220]

    bbox = find_best_bbox_on_page(page, probe)
    if bbox is None:
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat)
        return pix.tobytes("png"), None

    # 1) bbox 크기에 비례해 자동 패딩
    bw = bbox.width
    bh = bbox.height
    pad_x = max(margin, bw * pad_x_ratio)
    pad_y = max(margin, bh * pad_y_ratio)

    clip = fitz.Rect(bbox.x0 - pad_x, bbox.y0 - pad_y, bbox.x1 + pad_x, bbox.y1 + pad_y)

    # 2) 너무 좁게 잡힌 경우(줄 1개만 잡히는 등) → 페이지 폭에 가깝게 확장
    if clip.width < page_rect.width * min_w_ratio:
        # 좌우는 크게 열어주고, 위아래는 기존대로
        clip.x0 = page_rect.x0 + 10
        clip.x1 = page_rect.x1 - 10

    # 3) 높이가 너무 얇으면(한 줄만 잘려서 위/아래 잘리는 느낌) → 높이 보강
    if clip.height < min_h_px:
        mid = (clip.y0 + clip.y1) / 2
        clip.y0 = mid - (min_h_px / 2)
        clip.y1 = mid + (min_h_px / 2)

    # 4) 페이지 밖으로 나가지 않게 clamp
    clip = clip & page_rect

    mat = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=mat, clip=clip)
    return pix.tobytes("png"), clip


# =========================
# rules_stand 관련
# =========================
STOP_WORDS = {
    "알려줘","설명","설명해줘","뭐야","어떻게","왜","기준","규정","안전기준","표준","표준화",
    "적용","해당","관련","좀","주세요","가능","가능한","어떤","몇","어디","에서","으로","을","를","이","가"
}

def extract_keywords(question: str):
    q = (question or "").strip()
    q = re.sub(r"[^\w가-힣\s]", " ", q)
    tokens = [t.strip() for t in q.split() if t.strip()]

    kws = []
    for t in tokens:
        if len(t) < 2:
            continue
        if t in STOP_WORDS:
            continue
        kws.append(t)

    domain_boost = []
    if "면적" in q: domain_boost += ["면적", "유효면적", "㎡", "m2"]
    if "폭" in q: domain_boost += ["폭", "너비", "가로"]
    if "깊이" in q or "안길이" in q: domain_boost += ["깊이", "안길이", "세로"]
    if "문구" in q or "조항" in q: domain_boost += ["문구", "조항", "단서", "예외"]
    kws += domain_boost

    seen = set()
    out = []
    for k in kws:
        if k not in seen:
            seen.add(k)
            out.append(k)
    return out[:12]

def score_sentence(sent: str, keywords: list):
    s = sent.lower()
    score = 0
    hits = []

    for k in keywords:
        if k and k.lower() in s:
            score += 3
            hits.append(k)

    if re.search(r"\d+(\.\d+)?\s*(m|mm|cm|㎡|m2|M2)", sent):
        score += 2
        hits.append("치수/면적")

    if re.search(r"(이상|이하|초과|미만|이내)", sent):
        score += 1
        hits.append("기준조건")

    return score, hits

def highlight_text_md(text: str, keywords: list):
    if not text:
        return text
    out = text
    for k in sorted(set(keywords), key=len, reverse=True):
        if len(k) < 2:
            continue
        out = re.sub(re.escape(k), f"**{k}**", out, flags=re.IGNORECASE)
    return out

def extract_target_item_from_question(question: str, df_stand, top_k=3):
    q_raw = (question or "").strip()
    q = re.sub(r"\s+", "", q_raw).lower()

    if df_stand is None or "항목" not in df_stand.columns:
        return None, []

    items = df_stand["항목"].dropna().astype(str).unique().tolist()
    if not items:
        return None, []

    exact_hits = []
    for item in items:
        key = re.sub(r"\s+", "", item).lower()
        if key and key in q:
            exact_hits.append(item)

    core_hits = []
    if not exact_hits:
        for item in items:
            core = str(item).split("(")[0]
            core_key = re.sub(r"\s+", "", core).lower()
            if core_key and core_key in q:
                core_hits.append(item)

    strong = exact_hits or core_hits
    if strong:
        strong_sorted = sorted(strong, key=lambda x: len(str(x)), reverse=True)
        best = strong_sorted[0]
        cand = []
        for x in strong_sorted:
            if x not in cand:
                cand.append(x)
            if len(cand) >= top_k:
                break
        return best, cand

    def norm(s):
        return re.sub(r"\s+", "", str(s)).lower()

    qn = norm(q_raw)
    kws = extract_keywords(q_raw)
    if not kws:
        kws = [t for t in re.split(r"[^가-힣A-Za-z0-9]+", q_raw) if len(t) >= 2][:5]

    scored = []
    for item in items:
        it = norm(item)
        score = 0

        for k in kws:
            kn = norm(k)
            if kn and kn in it:
                score += 3

        if "장애인" in q_raw and "장애인" in item:
            score += 3
        if "면적" in q_raw and "면적" in item:
            score += 3
        if "폭" in q_raw and ("폭" in item or "너비" in item):
            score += 2
        if ("안길이" in q_raw or "깊이" in q_raw) and ("안길이" in item or "깊이" in item):
            score += 2

        if len(qn) >= 2 and qn in it:
            score += 2

        if score > 0:
            scored.append((score, item))

    scored.sort(key=lambda x: x[0], reverse=True)
    candidates = [it for _, it in scored[:top_k]]
    best = candidates[0] if candidates else None
    return best, candidates

def _to_date_safe(x):
    try:
        if pd.isna(x):
            return None
    except Exception:
        pass
    try:
        return pd.to_datetime(x, errors="coerce").date()
    except Exception:
        return None

def pick_applicable_row(df_stand: pd.DataFrame, target_item: str, base_date):
    if df_stand is None or df_stand.empty:
        return None

    df = df_stand.copy()
    df.columns = df.columns.astype(str).str.strip()
    if "항목" not in df.columns:
        return None

    subset = df[df["항목"] == target_item].copy()
    if subset.empty:
        return None

    start_col = "기준적용일"
    end_col = "기준전환일"
    if start_col not in subset.columns:
        return None

    subset["__start"] = subset[start_col].apply(_to_date_safe)
    subset["__end"] = subset[end_col].apply(_to_date_safe) if end_col in subset.columns else None

    if base_date is None:
        subset = subset.sort_values("__start", ascending=False)
        return subset.iloc[0].to_dict()

    def in_range(r):
        s = r["__start"]
        e = r["__end"]
        if s is not None and base_date < s:
            return False
        if e is not None and base_date >= e:
            return False
        return True

    app = subset[subset.apply(in_range, axis=1)].copy()
    if not app.empty:
        app = app.sort_values("__start", ascending=False)
        return app.iloc[0].to_dict()

    fallback = subset[subset["__start"].notna() & (subset["__start"] <= base_date)].copy()
    if not fallback.empty:
        fallback = fallback.sort_values("__start", ascending=False)
        return fallback.iloc[0].to_dict()

    subset = subset.sort_values("__start", ascending=False)
    return subset.iloc[0].to_dict()

# =========================
# rules 계산 결과 조회
# =========================
def get_rule_value(df_rules_calc: pd.DataFrame, item_name: str):
    if df_rules_calc is None or "항목" not in df_rules_calc.columns:
        return None
    tmp = df_rules_calc[df_rules_calc["항목"] == item_name]
    if tmp.empty:
        return None
    r = tmp.iloc[0]
    return {
        "항목": item_name,
        "입력값": r.get("입력값", None),
        "최소값": r.get("최소값", None),
        "최대값": r.get("최대값", None),
        "단위": r.get("단위", "") or "",
        "적합여부": r.get("적합여부", "") or "",
    }

STAND_TO_RULES_MAP = {
    "장애인용 엘리베이터 면적": ["장애인용 카치수(폭)", "장애인용 카치수(안길이)"],
}

def infer_group_items(target_item: str, df_rules_calc: pd.DataFrame):
    if target_item in STAND_TO_RULES_MAP:
        return STAND_TO_RULES_MAP[target_item]
    if df_rules_calc is None or "항목" not in df_rules_calc.columns:
        return [target_item]

    all_items = df_rules_calc["항목"].dropna().astype(str).tolist()
    prefix = target_item.split("(")[0].strip()
    group = [it for it in all_items if it.startswith(prefix)]
    return group if group else [target_item]

# =========================
# 세션 상태
# =========================
def init_session_state(df_input):
    if "input_values" not in st.session_state:
        st.session_state["input_values"] = {}
        for _, row in df_input.iterrows():
            key = row["입력키"]
            val = row["값"]
            if pd.isna(key):
                continue
            st.session_state["input_values"][key] = None if pd.isna(val) else val

    if "view" not in st.session_state:
        st.session_state["view"] = "입력"

def get_input_value(key):
    return st.session_state["input_values"].get(key, None)

def set_input_value(key, value):
    st.session_state["input_values"][key] = value

# =========================
# 입력 화면
# =========================
def render_input_page(df_input, cell_to_key, formula_key_to_formula):
    st.header("승강기 기본 제원 입력")
    st.write("※ 화면에는 '항목 / 값 / 단위'만 보여주고, 입력키는 내부적으로만 사용합니다.")

    for _, row in df_input.iterrows():
        item = row["항목"]
        key = row["입력키"]
        unit = row["단위"]
        raw_val = row["값"]

        if pd.isna(key):
            continue

        if key == "Standard application date":
            current = get_input_value(key)
            if isinstance(current, str) and current:
                try:
                    current = datetime.strptime(current, "%Y-%m-%d").date()
                except Exception:
                    current = None
            if not isinstance(current, date):
                current = date.today()

            new_val = st.date_input(
                f"{item} ({unit if isinstance(unit, str) else ''})",
                value=current,
                min_value=date(1995, 1, 1),
                max_value=date.today(),
            )
            set_input_value(key, new_val.isoformat())
            continue

        if key in formula_key_to_formula:
            value = evaluate_excel_formula(
                formula_key_to_formula[key],
                cell_to_key,
                st.session_state["input_values"],
            )
            if value is not None:
                set_input_value(key, value)
                vnum = numeric_or_none(value)
                disp = f"{vnum:.2f}" if vnum is not None else str(value)
                st.text_input(f"{item} (자동 계산, {unit})", value=disp, disabled=True)
            else:
                st.text_input(f"{item} (자동 계산, {unit})", value="입력값 부족 / 계산불가", disabled=True)
            continue

        if isinstance(unit, str) and unit.strip() == "선택" and isinstance(raw_val, str):
            options = [opt.strip() for opt in raw_val.split(",")]
            current = get_input_value(key)
            if current not in options:
                current = options[0] if options else ""
            new_val = st.selectbox(
                f"{item}",
                options=options,
                index=options.index(current) if current in options else 0,
            )
            set_input_value(key, new_val)
            st.caption(f"단위: {unit}")
            continue

        current = get_input_value(key)
        if current is None or (isinstance(current, float) and math.isnan(current)):
            current = "" if pd.isna(raw_val) else str(raw_val)

        new_val = st.text_input(f"{item} ({unit if isinstance(unit, str) else ''})", value=str(current))
        set_input_value(key, new_val)

    if st.button("입력 완료 후 규칙 확인 화면으로 이동"):
        st.session_state["view"] = "규칙"
        st.rerun()

# =========================
# Q1: rules 기반 Q&A (검색형 / 판정 없음)
# =========================
def get_stand_text_from_row(row: dict) -> str:
    if not row:
        return ""
    candidates = ["기준내용","비고","기준 내용","비 고","기준설명","기준 설명","설명","내용"]
    for k in candidates:
        v = row.get(k, None)
        if v is None:
            continue
        s = str(v).strip()
        if s and s.lower() not in ["nan", "none"]:
            return s
    for k, v in row.items():
        if not isinstance(k, str):
            continue
        kk = k.replace(" ", "")
        if ("기준" in kk or "내용" in kk or "비고" in kk) and v is not None:
            s = str(v).strip()
            if s and s.lower() not in ["nan", "none"]:
                return s
    return ""

def strip_time(v):
    if not v:
        return ""
    s = str(v)
    if s.lower() == "nat":
        return "현재"
    return s.split(" ")[0]

def answer_rules_based_question_search(question, target_item_sel, df_rules_calc, df_stand):
    base_date_raw = get_input_value("Standard application date")
    base_date = pd.to_datetime(base_date_raw).date() if base_date_raw else None

    if target_item_sel and target_item_sel != "(질문에서 자동 인식)":
        target_item = target_item_sel
        candidates = []
    else:
        target_item, candidates = extract_target_item_from_question(question, df_stand)

    if not target_item:
        msg = (
            "### 결과\n"
            "- 질문에서 기준 항목을 특정하지 못했습니다.\n"
            "- 드롭다운에서 기준 항목을 선택해 주세요.\n"
        )
        if candidates:
            msg += "\n**추천 후보(Top3):**\n" + "\n".join([f"- {c}" for c in candidates])
        return msg

    applicable = pick_applicable_row(df_stand, target_item, base_date)

    if not applicable:
        subset = df_stand[df_stand["항목"] == target_item].copy()
        if not subset.empty:
            subset["__start"] = pd.to_datetime(subset.get("기준적용일"), errors="coerce")
            subset = subset.sort_values("__start", ascending=False)
            applicable = subset.iloc[0].to_dict()

    stand_text = ""
    s = e = ""
    if applicable:
        s = strip_time(applicable.get("기준적용일", "") or "")
        e = strip_time(
            applicable.get("기준전환일", "")
            or applicable.get("반환일", "")
            or applicable.get("기준종료일", "")
            or ""
        )
        stand_text = get_stand_text_from_row(applicable) or ""

    keywords = extract_keywords(question)
    sents = split_sentences_ko(stand_text)

    scored = []
    for sent in sents:
        sc, _hits = score_sentence(sent, keywords)
        if sc > 0:
            scored.append((sc, sent))
    scored.sort(key=lambda x: x[0], reverse=True)
    top = scored[:6]

    lines = []
    lines.append("### 기준 적용 Q&A (검색형 · 판정 없음)")
    lines.append(f"- **현장 기준적용일:** {base_date if base_date else '미입력'}")
    lines.append(f"- **해당 기준 항목:** {target_item}")

    lines.append("")
    lines.append("#### 1) 적용기간")
    if applicable and (s or e):
        lines.append(f"- {s} ~ {e}")
    else:
        lines.append("- rules_stand에서 기간 정보를 찾지 못했습니다(또는 비어 있음).")

    lines.append("")
    lines.append("#### 2) 기준 내용(질문 키워드 기반 발췌)")
    if not stand_text.strip():
        lines.append("- (rules_stand 기준내용/비고가 비어 있습니다.)")
    elif not top:
        lines.append("- (질문 키워드와 직접 매칭되는 문장을 찾지 못했습니다.)")
        lines.append("- 아래 ‘전체 기준내용(원문)’을 참고하세요.")
    else:
        for i, (_, sent) in enumerate(top, 1):
            lines.append(f"- **[{i}]** {highlight_text_md(sent, keywords)}")

    if stand_text.strip():
        lines.append("")
        lines.append("#### 3) 전체 기준내용(원문)")
        lines.append(
            "<details><summary>펼쳐보기</summary>\n\n"
            + highlight_text_md(short_quote(stand_text, 2500), keywords)
            + "\n\n</details>"
        )

    rule_items = infer_group_items(target_item, df_rules_calc)
    ref_lines = []
    for it in rule_items:
        rv = get_rule_value(df_rules_calc, it)
        if not rv:
            continue
        mn0 = numeric_or_none(rv.get("최소값"))
        mx0 = numeric_or_none(rv.get("최대값"))
        unit = rv.get("단위", "") or ""
        if mn0 is None and mx0 is None:
            continue
        if mn0 is not None and mx0 is not None:
            ref_lines.append(f"- {it}: **{mn0:g} ~ {mx0:g}{unit}**")
        elif mn0 is not None:
            ref_lines.append(f"- {it}: **최소 {mn0:g}{unit} 이상**")
        else:
            ref_lines.append(f"- {it}: **최대 {mx0:g}{unit} 이하**")

    if ref_lines:
        lines.append("")
        lines.append("#### 4) 수치 기준(참고: rules 시트)")
        lines.extend(ref_lines)

    lines.append("")
    lines.append("> ※ Q1은 **기준 설명/발췌** 전용이며, 적합/부적합 **판정은 제공하지 않습니다.**")
    return "\n".join(lines)

# =========================
# rules 화면 (표 + Q1/Q2)
# =========================
def render_rules_page(df_rules, df_stand, input_cell_to_key):
    st.header("기준 적합성 확인")

    df = df_rules.copy().reset_index(drop=True)
    input_map = st.session_state.get("input_values", {})

    formulas_min, formulas_max, formulas_val = load_rules_formulas()
    bs_values = load_bs_values()
    bs_tables = load_bs_tables()

    filled_values, judgements, min_list, max_list = [], [], [], []

    for idx, row in df.iterrows():
        key = row.get("입력키", None)
        item_name = row.get("항목", "")
        min_v = row.get("최소값", None)
        max_v = row.get("최대값", None)
        val_v = row.get("입력값", None)

        if idx in formulas_min:
            calc_min = evaluate_rules_formula(formulas_min[idx], input_cell_to_key, input_map, bs_values)
            if calc_min is not None:
                min_v = calc_min
        if idx in formulas_max:
            calc_max = evaluate_rules_formula(formulas_max[idx], input_cell_to_key, input_map, bs_values)
            if calc_max is not None:
                max_v = calc_max
        if idx in formulas_val:
            calc_val = evaluate_rules_formula(formulas_val[idx], input_cell_to_key, input_map, bs_values)
            if calc_val is not None:
                val_v = calc_val

        if isinstance(key, str) and key in input_map:
            val_v = input_map.get(key, val_v)

        if item_name == "카 유효면적":
            ca_min = compute_car_area_min(input_map, bs_tables)
            ca_max = compute_car_area_max(input_map, bs_tables)
            if ca_min is not None:
                min_v = ca_min
            if ca_max is not None:
                max_v = ca_max

        min_list.append(min_v)
        max_list.append(max_v)
        filled_values.append(val_v)
        judgements.append(judge_conform(val_v, min_v, max_v))

    df["최소값"] = min_list
    df["최대값"] = max_list
    df["입력값"] = filled_values
    df["적합여부"] = judgements
    st.session_state["rules_calc_df"] = df.copy()

    st.subheader("항목별 최소/최대값 대비 입력값 적합 여부")
    show_df = df[["항목", "최소값", "최대값", "단위", "입력값", "적합여부"]].copy()

    def fmt2(x):
        v = numeric_or_none(x)
        return "" if v is None else f"{v:.2f}"

    for col in ["최소값", "최대값", "입력값"]:
        if col in show_df.columns:
            show_df[col] = show_df[col].apply(fmt2)

    def style_conform_row(row):
        j = str(row.get("적합여부", "")).strip()
        if j == "적합":
            return ["background-color: #d7f5d7"] * len(row)
        if j == "부적합":
            return ["background-color: #ffd6d6"] * len(row)
        return [""] * len(row)

    styled = show_df.style.apply(style_conform_row, axis=1)
    st.dataframe(styled, use_container_width=True)

    # ---------------------------
    # Q1
    # ---------------------------
    st.markdown("---")
    st.subheader("기준적용일·표준화에 따른 안전기준 적용 Q&A")

    with st.form("rules_based_question_form"):
        q1 = st.text_input("기준 관련 질문을 입력하세요 (예: 장애인용 면적 알려줘)", key="q1")

        item_options = ["(질문에서 자동 인식)"] + sorted(df_stand["항목"].dropna().unique().tolist())
        target_item_sel = st.selectbox(
            "해당 기준 항목 (선택하면 더 정확하게 찾을 수 있습니다)",
            options=item_options,
            key="q1_item",
        )
        submitted1 = st.form_submit_button("기준 답변")

    if submitted1:
        ans1 = answer_rules_based_question_search(
            q1,
            target_item_sel,
            st.session_state.get("rules_calc_df"),
            df_stand,
        )
        st.markdown(ans1, unsafe_allow_html=True)

    # ---------------------------
    # ✅ Q2 (반드시 render_rules_page 내부!)
    # ---------------------------
    st.markdown("---")
    st.subheader("질문 2) GPT 연동(연혁·표준화 내용 질의)")

    with st.form("pdf_gpt_question_form"):
        q2 = st.text_input("연혁/표준화 PDF에 대해 질문하세요", key="q2")
        top_k = st.slider("참고 문단 개수", min_value=1, max_value=5, value=3)
        use_gpt = st.checkbox("GPT로 요약/정리까지 받기", value=True)
        submitted2 = st.form_submit_button("답변")

    if submitted2:
        if not q2.strip():
            st.warning("질문을 입력해 주세요.")
        else:
            hits = search_local_standard(q2, top_k=top_k)

            if not hits:
                st.write("결과: (PDF 지식베이스에서 관련 문단을 찾지 못했습니다)")
            else:
                st.write("### PDF 발췌(근거)")
                for i, h in enumerate(hits, 1):
                    src = h.get("source", "문서")
                    pg = h.get("page", "")
                    txt_full = h.get("text", "")

                    st.markdown(f"**[{i}] {src} (p.{pg})**")
                    st.write(short_quote(txt_full, 900))

                    pdf_path = resolve_pdf_path(src)
                    if not pdf_path:
                        st.warning(f"PDF 캡처 실패: 경로를 찾지 못함 (source={src})")
                        continue
                    if not str(pg).strip().isdigit():
                        st.warning(f"PDF 캡처 실패: 페이지 정보 이상 (page={pg})")
                        continue

                    try:
                        png_bytes, _ = render_pdf_snippet(
                            pdf_path=pdf_path,
                            page_no_1based=int(pg),
                            quote=txt_full,
                            margin=14,
                            zoom=2.0,
                        )
                        st.image(
                            png_bytes,
                            caption=f"PDF 캡처: {os.path.basename(pdf_path)} (p.{pg})",
                            use_container_width=True,
                        )
                    except Exception as e:
                        st.warning(f"PDF 캡처 실패: {e}")

                if use_gpt:
                    context = "\n\n".join(
                        [f"[{h.get('source','문서')} p.{h.get('page','')}] {h.get('text','')}" for h in hits]
                    )
                    gpt_ans = ask_gpt_about_standard(q2, context_text=context)
                    st.markdown("---")
                    st.markdown("### GPT 정리(요약/해석)")
                    st.markdown(gpt_ans)

# =========================
# 헤더
# =========================
def render_header():
    cols = st.columns([0.6, 4])
    with cols[0]:
        if os.path.exists(EMBLEM_PATH):
            st.image(EMBLEM_PATH, width=120)
    with cols[1]:
        st.markdown(
            """
            <div style='display:flex; flex-direction:column; justify-content:center;'>
              <h1 style='margin-bottom:4px;'>승강기 정격·기준 자동 체크 시스템</h1>
              <p style='margin-top:0; color:#555;'>한국승강기안전공단 안전인증실 허동민 과장</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    st.markdown("---")

# =========================
# 메인
# =========================
def main():
    st.set_page_config(
        page_title="승강기 정격/기준 체크 도구",
        layout="wide",
        page_icon=EMBLEM_PATH if os.path.exists(EMBLEM_PATH) else None,
    )

    if not EXCEL_PATH.exists():
        st.error(f"엑셀 파일을 찾을 수 없습니다: {EXCEL_PATH}")
        st.stop()

    render_header()

    df_input, df_rules, df_stand = load_excel_frames()
    df_stand = preprocess_df_stand(df_stand)
    cell_to_key, formula_key_to_formula = load_input_structure()
    init_session_state(df_input)

    view = st.sidebar.radio(
        "화면 선택",
        ["입력", "규칙"],
        index=0 if st.session_state["view"] == "입력" else 1,
    )
    st.session_state["view"] = view

    if view == "입력":
        render_input_page(df_input, cell_to_key, formula_key_to_formula)
    else:
        render_rules_page(df_rules, df_stand, cell_to_key)

if __name__ == "__main__":
    main()

